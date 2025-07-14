import streamlit as st
import pandas as pd
from tabula import read_pdf
import pdfplumber
from PIL import Image
import io
import os
import tempfile
import zipfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage


def extract_tables_from_pdf(pdf_path):
    """Extract tables from PDF using tabula"""
    try:
        tables = read_pdf(pdf_path, pages="all", multiple_tables=True)
        return tables
    except Exception as e:
        st.error(f"Error extracting tables: {str(e)}")
        return []


def extract_images_from_pdf(pdf_path):
    """Extract images from PDF using pdfplumber"""
    images = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Get images from the page
                if hasattr(page, 'images') and page.images:
                    for img_index, img_obj in enumerate(page.images):
                        try:
                            # Extract image properties
                            x0, y0, x1, y1 = img_obj['x0'], img_obj['y0'], img_obj['x1'], img_obj['y1']

                            # Crop the image from the page
                            cropped_page = page.crop((x0, y0, x1, y1))

                            # Convert to PIL Image
                            img_pil = cropped_page.to_image(resolution=150)

                            # Convert PIL image to bytes
                            img_bytes = io.BytesIO()
                            img_pil.save(img_bytes, format='PNG')
                            img_bytes = img_bytes.getvalue()

                            images.append({
                                'page': page_num + 1,
                                'index': img_index,
                                'bytes': img_bytes,
                                'ext': 'png',
                                'filename': f"image_page{page_num + 1}_{img_index}.png"
                            })
                        except Exception as img_error:
                            st.warning(
                                f"Could not extract image {img_index} from page {page_num + 1}: {str(img_error)}")
                            continue

                # Alternative method: try to extract figures/charts
                try:
                    # Convert entire page to image if no specific images found
                    if not (hasattr(page, 'images') and page.images):
                        # Check if page has visual content by looking for drawings/figures
                        if page.figures or page.curves or page.rects:
                            # Create image from page
                            img_pil = page.to_image(resolution=150)

                            # Convert PIL image to bytes
                            img_bytes = io.BytesIO()
                            img_pil.save(img_bytes, format='PNG')
                            img_bytes = img_bytes.getvalue()

                            images.append({
                                'page': page_num + 1,
                                'index': 0,
                                'bytes': img_bytes,
                                'ext': 'png',
                                'filename': f"page_{page_num + 1}_visual_content.png"
                            })
                except Exception as page_error:
                    continue

        return images
    except Exception as e:
        st.error(f"Error extracting images: {str(e)}")
        return []


def extract_images_alternative_method(pdf_path):
    """Alternative method to extract images by converting pages to images"""
    images = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                try:
                    # Convert page to image
                    img_pil = page.to_image(resolution=150)

                    # Convert PIL image to bytes
                    img_bytes = io.BytesIO()
                    img_pil.save(img_bytes, format='PNG')
                    img_bytes = img_bytes.getvalue()

                    images.append({
                        'page': page_num + 1,
                        'index': 0,
                        'bytes': img_bytes,
                        'ext': 'png',
                        'filename': f"page_{page_num + 1}.png"
                    })
                except Exception as page_error:
                    st.warning(
                        f"Could not convert page {page_num + 1} to image: {str(page_error)}")
                    continue

        return images
    except Exception as e:
        st.error(f"Error in alternative image extraction: {str(e)}")
        return []


def create_excel_with_tables_and_images(tables, images):
    """Create Excel file with tables and images"""
    # Create temporary file for Excel
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        excel_path = tmp_file.name

    # Create Excel file with tables
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        if tables:
            for i, table in enumerate(tables):
                sheet_name = f"Table_{i+1}"
                table.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # Create empty sheet if no tables found
            pd.DataFrame({'Message': ['No tables found in PDF']}).to_excel(
                writer, sheet_name='No_Tables', index=False
            )

    # Add images to Excel if any exist
    if images:
        wb = load_workbook(excel_path)
        ws = wb.create_sheet(title="Extracted_Images")

        # Create temporary directory for images
        with tempfile.TemporaryDirectory() as temp_dir:
            row = 1
            for img_data in images:
                # Save image temporarily
                img_path = os.path.join(temp_dir, img_data['filename'])
                with open(img_path, "wb") as f:
                    f.write(img_data['bytes'])

                # Add image to Excel
                try:
                    excel_img = ExcelImage(img_path)
                    # Resize image to fit better in Excel
                    excel_img.width = min(excel_img.width, 400)
                    excel_img.height = min(excel_img.height, 300)
                    ws.add_image(excel_img, f"A{row}")

                    # Add image info in adjacent columns
                    ws[f"E{row}"] = f"Page {img_data['page']}, Image {img_data['index'] + 1}"
                    ws[f"E{row + 1}"] = img_data['filename']

                    row += 20  # Space between images
                except Exception as e:
                    st.warning(
                        f"Could not add image {img_data['filename']} to Excel: {str(e)}")

        wb.save(excel_path)

    return excel_path


def create_images_zip(images):
    """Create ZIP file with all extracted images"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp_file:
        zip_path = tmp_file.name

    with zipfile.ZipFile(zip_path, 'w') as zip_file:
        for img_data in images:
            zip_file.writestr(img_data['filename'], img_data['bytes'])

    return zip_path


def main():
    st.set_page_config(
        page_title="PDF Table & Image Extractor",
        page_icon="📄",
        layout="wide"
    )

    st.title("📄 PDF Table & Image Extractor")
    st.markdown(
        "Upload a PDF file to extract tables and images into downloadable formats.")

    # File upload
    uploaded_file = st.file_uploader(
        "Choose a PDF file",
        type="pdf",
        help="Upload a PDF file to extract tables and images"
    )

    if uploaded_file is not None:
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            pdf_path = tmp_file.name

        st.success(f"✅ PDF uploaded successfully: {uploaded_file.name}")

        # Processing options
        col1, col2, col3 = st.columns(3)
        with col1:
            extract_tables = st.checkbox("Extract Tables", value=True)
        with col2:
            extract_images = st.checkbox("Extract Images", value=True)
        with col3:
            extract_pages = st.checkbox("Extract Pages as Images", value=False,
                                        help="Convert entire pages to images")

        if st.button("🚀 Start Extraction", type="primary"):
            if not extract_tables and not extract_images and not extract_pages:
                st.warning("Please select at least one extraction option.")
                return

            with st.spinner("Processing PDF..."):
                tables = []
                images = []

                # Extract tables
                if extract_tables:
                    st.info("Extracting tables...")
                    tables = extract_tables_from_pdf(pdf_path)
                    if tables:
                        st.success(f"✅ Found {len(tables)} tables")
                    else:
                        st.warning("⚠️ No tables found")

                # Extract images
                if extract_images:
                    st.info("Extracting images...")
                    images = extract_images_from_pdf(pdf_path)
                    if images:
                        st.success(f"✅ Found {len(images)} images")
                    else:
                        st.warning("⚠️ No embedded images found")

                # Extract pages as images
                if extract_pages:
                    st.info("Converting pages to images...")
                    page_images = extract_images_alternative_method(pdf_path)
                    if page_images:
                        images.extend(page_images)
                        st.success(
                            f"✅ Converted {len(page_images)} pages to images")
                    else:
                        st.warning("⚠️ Could not convert pages to images")

                # Display results
                if tables or images:
                    st.subheader("📊 Results")

                    # Show table previews
                    if tables:
                        st.subheader("📋 Extracted Tables")
                        for i, table in enumerate(tables):
                            with st.expander(f"Table {i+1} (Shape: {table.shape})"):
                                st.dataframe(table)

                    # Show image previews
                    if images:
                        st.subheader("🖼️ Extracted Images")
                        cols = st.columns(3)
                        for i, img_data in enumerate(images):
                            with cols[i % 3]:
                                try:
                                    img = Image.open(
                                        io.BytesIO(img_data['bytes']))
                                    st.image(
                                        img, caption=f"Page {img_data['page']}", use_column_width=True)
                                except Exception as e:
                                    st.error(
                                        f"Could not display image: {str(e)}")

                    # Download options
                    st.subheader("📥 Download Options")
                    col1, col2 = st.columns(2)

                    with col1:
                        if tables or images:
                            st.info(
                                "Creating Excel file with tables and images...")
                            excel_path = create_excel_with_tables_and_images(
                                tables, images)

                            with open(excel_path, "rb") as file:
                                st.download_button(
                                    label="📊 Download Excel File",
                                    data=file.read(),
                                    file_name=f"{uploaded_file.name.rsplit('.', 1)[0]}_extracted.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                            # Clean up temporary Excel file
                            os.unlink(excel_path)

                    with col2:
                        if images:
                            st.info("Creating ZIP file with images...")
                            zip_path = create_images_zip(images)

                            with open(zip_path, "rb") as file:
                                st.download_button(
                                    label="🖼️ Download Images ZIP",
                                    data=file.read(),
                                    file_name=f"{uploaded_file.name.rsplit('.', 1)[0]}_images.zip",
                                    mime="application/zip"
                                )

                            # Clean up temporary ZIP file
                            os.unlink(zip_path)
                else:
                    st.error("❌ No tables or images found in the PDF")

        # Clean up temporary PDF file
        os.unlink(pdf_path)

    # Instructions
    with st.expander("📖 Instructions"):
        st.markdown("""
        ### How to use this tool:
        1. **Upload PDF**: Click "Choose a PDF file" and select your PDF document
        2. **Select Options**: Choose extraction options:
           - **Extract Tables**: Find and extract data tables
           - **Extract Images**: Extract embedded images from PDF
           - **Extract Pages as Images**: Convert entire pages to PNG images
        3. **Start Extraction**: Click "Start Extraction" to process the PDF
        4. **Preview Results**: Review the extracted tables and images
        5. **Download**: Use the download buttons to get:
           - Excel file with tables and images
           - ZIP file with all extracted images
        
        ### Features:
        - ✅ Extract tables from all pages using tabula-py
        - ✅ Extract embedded images using pdfplumber
        - ✅ Convert entire pages to images (useful for charts/graphics)
        - ✅ Preview results before downloading
        - ✅ Excel file with separate sheets for tables and images
        - ✅ ZIP file with all images for easy access
        - ✅ Automatic file naming based on original PDF
        - ✅ Cloud-compatible (no Java dependencies for images)
        
        ### Tips:
        - If no embedded images are found, try "Extract Pages as Images"
        - For PDFs with complex layouts, page extraction might work better
        - Table extraction works best with structured tabular data
        
        ### Requirements:
        This app uses pdfplumber for reliable image extraction on cloud platforms.
        """)


if __name__ == "__main__":
    main()
